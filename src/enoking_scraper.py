"""Public Enoking buyback product scraper.

The scraper only reads public pages, runs at a low default frequency, and does not
log in, bypass access controls, add items to carts, or automate purchases.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin, urlparse
from urllib.robotparser import RobotFileParser

import requests
from bs4 import BeautifulSoup, Tag

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
JST = timezone(timedelta(hours=9), "JST")

DEFAULT_BASE_URL = "https://newenoking-kaitori.com"
USER_AGENT = os.getenv(
    "ENOKING_USER_AGENT",
    "Mozilla/5.0 (compatible; EnokingCloudMonitor/2.0; +https://github.com/univcorp2-ctrl/enoking-monitor-starter)",
)
REQUEST_TIMEOUT_SEC = int(os.getenv("REQUEST_TIMEOUT_SEC", "25"))
REQUEST_INTERVAL_SEC = float(os.getenv("REQUEST_INTERVAL_SEC", "1.0"))
MAX_CRAWL_PAGES = int(os.getenv("ENOKING_MAX_CRAWL_PAGES", "220"))
RESPECT_ROBOTS = os.getenv("RESPECT_ROBOTS", "true").lower() not in {"0", "false", "no"}

COMMON_NON_NOTES = {
    "☆",
    "備考 ▼",
    "参考買取金額",
    "店舗ごとの買取金額を見る 現金買取可能",
    "店舗ごとの買取金額を見る",
    "現金買取可能",
    "商品数",
    "−＋",
    "カートに追加",
}


@dataclass(frozen=True)
class CategoryLink:
    name: str
    url: str


@dataclass
class EnokingProduct:
    jan: str
    product_name: str
    category: str
    buy_price_yen: int
    currency: str
    notes: str
    cash_purchase_available: bool
    image_url: str
    source_url: str
    is_featured: bool
    scraped_at_jst: str
    source_hash: str


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\u3000", " ")).strip()


def normalize_lines(value: str) -> list[str]:
    return [normalize_space(line) for line in value.splitlines() if normalize_space(line)]


def parse_yen(value: str) -> int | None:
    digits = re.sub(r"[^0-9]", "", value)
    if not digits:
        return None
    return int(digits)


def extract_jan(text: str) -> str | None:
    match = re.search(r"JAN\s*[:：]\s*([0-9]{8,14})", text)
    return match.group(1) if match else None


def text_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8", errors="ignore")).hexdigest()[:16]


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
    )
    return session


def build_robot_parser(base_url: str, session: requests.Session) -> RobotFileParser | None:
    if not RESPECT_ROBOTS:
        return None
    robots_url = urljoin(base_url, "/robots.txt")
    parser = RobotFileParser()
    parser.set_url(robots_url)
    try:
        response = session.get(robots_url, timeout=REQUEST_TIMEOUT_SEC)
        if response.status_code >= 400 or not response.text.strip():
            return None
        parser.parse(response.text.splitlines())
        return parser
    except requests.RequestException:
        return None


def robot_allows(parser: RobotFileParser | None, url: str) -> bool:
    if parser is None:
        return True
    return parser.can_fetch(USER_AGENT, url)


def fetch_html(session: requests.Session, url: str, robot_parser: RobotFileParser | None = None) -> str:
    if not robot_allows(robot_parser, url):
        raise RuntimeError(f"robots.txt disallows fetch: {url}")
    response = session.get(url, timeout=REQUEST_TIMEOUT_SEC)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or response.encoding
    return response.text


def extract_category_links(html: str, base_url: str = DEFAULT_BASE_URL) -> list[CategoryLink]:
    soup = BeautifulSoup(html, "html.parser")
    links: list[CategoryLink] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href"))
        label = normalize_space(anchor.get_text(" ", strip=True))
        if not label or "cat=" not in href or "/products" not in href:
            continue
        url = urljoin(base_url, href)
        if url in seen:
            continue
        seen.add(url)
        links.append(CategoryLink(name=label.replace("▸", "").strip(), url=url))
    return links


def is_probable_product_heading(tag: Tag) -> bool:
    if tag.name not in {"h2", "h3"}:
        return False
    text = normalize_space(tag.get_text(" ", strip=True))
    if not text:
        return False
    parent_text = normalize_space(tag.parent.get_text(" ", strip=True) if tag.parent else "")
    return "JAN" in parent_text and "参考買取金額" in parent_text


def smallest_product_container(heading: Tag) -> Tag:
    best = heading.parent or heading
    for parent in [heading.parent, *list(heading.parents)]:
        if not isinstance(parent, Tag):
            continue
        text = normalize_space(parent.get_text(" ", strip=True))
        if "JAN" in text and "参考買取金額" in text and len(text) < 1800:
            best = parent
            break
    return best


def infer_category(lines: list[str], product_name: str, fallback_category: str) -> str:
    try:
        idx = lines.index(product_name)
    except ValueError:
        return fallback_category
    for candidate in reversed(lines[:idx]):
        if candidate in COMMON_NON_NOTES:
            continue
        if candidate.startswith("JAN") or candidate.startswith("¥"):
            continue
        if candidate == product_name:
            continue
        return candidate
    return fallback_category


def infer_notes(lines: list[str]) -> str:
    notes: list[str] = []
    in_price_area = False
    for line in lines:
        if line == "参考買取金額":
            in_price_area = True
            continue
        if line == "商品数":
            break
        if not in_price_area:
            continue
        if line.startswith("¥") or line in COMMON_NON_NOTES or line.startswith("JAN"):
            continue
        notes.append(line)
    return " / ".join(dict.fromkeys(notes))


def find_image_url(container: Tag, base_url: str) -> str:
    image = container.find("img")
    if not isinstance(image, Tag):
        return ""
    src = str(image.get("src") or image.get("data-src") or "").strip()
    return urljoin(base_url, src) if src else ""


def parse_products_from_html(
    html: str,
    source_url: str,
    fallback_category: str = "",
    scraped_at_jst: str | None = None,
) -> list[EnokingProduct]:
    soup = BeautifulSoup(html, "html.parser")
    base_url = f"{urlparse(source_url).scheme}://{urlparse(source_url).netloc}" if urlparse(source_url).netloc else DEFAULT_BASE_URL
    scraped_at = scraped_at_jst or datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S %Z")
    products: list[EnokingProduct] = []

    for heading in soup.find_all(is_probable_product_heading):
        container = smallest_product_container(heading)
        raw_text = container.get_text("\n", strip=True)
        lines = normalize_lines(raw_text)
        combined = "\n".join(lines)
        jan = extract_jan(combined)
        price_match = re.search(r"参考買取金額\s*\n?\s*¥?\s*([0-9,]+)", combined)
        product_name = normalize_space(heading.get_text(" ", strip=True))
        price = parse_yen(price_match.group(1)) if price_match else None
        if not jan or not product_name or price is None:
            continue
        category = infer_category(lines, product_name, fallback_category)
        products.append(
            EnokingProduct(
                jan=jan,
                product_name=product_name,
                category=category,
                buy_price_yen=price,
                currency="JPY",
                notes=infer_notes(lines),
                cash_purchase_available="現金買取可能" in combined,
                image_url=find_image_url(container, base_url),
                source_url=source_url,
                is_featured="/featured" in source_url,
                scraped_at_jst=scraped_at,
                source_hash=text_hash(raw_text),
            )
        )
    return products


def pagination_links(html: str, current_url: str, base_url: str = DEFAULT_BASE_URL) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    links: list[str] = []
    for anchor in soup.find_all("a", href=True):
        label = normalize_space(anchor.get_text(" ", strip=True))
        href = str(anchor.get("href"))
        if not href:
            continue
        if label == "次へ ›" or label == "次へ" or re.fullmatch(r"[0-9]+", label):
            url = urljoin(base_url, href)
            if url != current_url and "/products" in url or "/featured" in url:
                links.append(url)
    return list(dict.fromkeys(links))


def dedupe_products(products: Iterable[EnokingProduct]) -> list[EnokingProduct]:
    deduped: dict[str, EnokingProduct] = {}
    for product in products:
        existing = deduped.get(product.jan)
        if existing is None:
            deduped[product.jan] = product
            continue
        existing_sources = set(filter(None, existing.source_url.split(" | ")))
        existing_sources.add(product.source_url)
        existing.source_url = " | ".join(sorted(existing_sources))
        existing.is_featured = existing.is_featured or product.is_featured
        if product.category and product.category not in existing.category.split(" / "):
            existing.category = f"{existing.category} / {product.category}" if existing.category else product.category
        if product.buy_price_yen != existing.buy_price_yen:
            existing.notes = normalize_space(f"{existing.notes} / price_changed_or_multi_page_seen").strip(" /")
            existing.buy_price_yen = max(existing.buy_price_yen, product.buy_price_yen)
    return sorted(deduped.values(), key=lambda item: (item.category, item.product_name, item.jan))


def crawl_enoking_products(base_url: str = DEFAULT_BASE_URL) -> list[EnokingProduct]:
    session = make_session()
    robot_parser = build_robot_parser(base_url, session)
    products_url = urljoin(base_url, "/products")
    featured_url = urljoin(base_url, "/featured")
    visited: set[str] = set()
    queue: list[tuple[str, str]] = [(featured_url, "注目商品"), (products_url, "")]
    scraped: list[EnokingProduct] = []
    scraped_at = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S %Z")

    while queue and len(visited) < MAX_CRAWL_PAGES:
        url, category_name = queue.pop(0)
        if url in visited:
            continue
        visited.add(url)
        html = fetch_html(session, url, robot_parser)
        scraped.extend(parse_products_from_html(html, url, category_name, scraped_at))

        if url == products_url:
            for category in extract_category_links(html, base_url):
                if category.url not in visited:
                    queue.append((category.url, category.name))

        for page_url in pagination_links(html, url, base_url):
            if page_url not in visited:
                queue.append((page_url, category_name))

        if REQUEST_INTERVAL_SEC > 0 and queue:
            time.sleep(REQUEST_INTERVAL_SEC)

    return dedupe_products(scraped)


def product_to_dict(product: EnokingProduct) -> dict[str, object]:
    return asdict(product)


def write_csv(products: list[EnokingProduct], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(asdict(products[0]).keys()) if products else ["message"]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        if products:
            writer.writerows(product_to_dict(product) for product in products)
        else:
            writer.writerow({"message": "NO_PRODUCTS"})


def write_json(products: list[EnokingProduct], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump([product_to_dict(product) for product in products], file, ensure_ascii=False, indent=2)


def autosize_sheet(ws: object) -> None:
    if Workbook is None:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in ws.columns:
        max_len = 0
        for cell in column:
            max_len = max(max_len, min(len("" if cell.value is None else str(cell.value)), 72))
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(10, max_len + 2)


def write_xlsx(products: list[EnokingProduct], path: Path) -> None:
    if Workbook is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "EnokingProducts"
    fieldnames = list(asdict(products[0]).keys()) if products else ["message"]
    ws.append(fieldnames)
    if products:
        for product in products:
            ws.append([product_to_dict(product).get(field) for field in fieldnames])
    else:
        ws.append(["NO_PRODUCTS"])
    autosize_sheet(ws)
    wb.save(path)


def write_outputs(products: list[EnokingProduct], output_dir: Path = OUTPUT_DIR) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    paths = {
        "csv": output_dir / f"enoking_products_{stamp}.csv",
        "json": output_dir / f"enoking_products_{stamp}.json",
        "xlsx": output_dir / f"enoking_products_{stamp}.xlsx",
        "latest_csv": output_dir / "latest_enoking_products.csv",
        "latest_json": output_dir / "latest_enoking_products.json",
        "latest_xlsx": output_dir / "latest_enoking_products.xlsx",
    }
    write_csv(products, paths["csv"])
    write_json(products, paths["json"])
    write_xlsx(products, paths["xlsx"])
    write_csv(products, paths["latest_csv"])
    write_json(products, paths["latest_json"])
    write_xlsx(products, paths["latest_xlsx"])
    return {key: str(value) for key, value in paths.items()}


def main() -> int:
    products = crawl_enoking_products(os.getenv("ENOKING_BASE_URL", DEFAULT_BASE_URL))
    paths = write_outputs(products)
    summary = {
        "scraped_at_jst": datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S %Z"),
        "product_count": len(products),
        "outputs": paths,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
