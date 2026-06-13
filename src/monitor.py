"""Low-load supplier monitor for Enoking buyback checks.

The monitor fetches configured supplier pages, extracts conservative price and
stock signals, compares them with Enoking buyback prices, and writes CSV/XLSX
reports for human review.

It does not automate purchases, login, CAPTCHA handling, waiting-room bypass,
cart actions, or order actions.
"""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import sqlite3
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "config"
OUTPUT_DIR = ROOT / "output"
JST = timezone(timedelta(hours=9), "JST")

BUY_MARGIN_THRESHOLD_YEN = int(os.getenv("BUY_MARGIN_THRESHOLD_YEN", "2000"))
REQUEST_TIMEOUT_SEC = int(os.getenv("REQUEST_TIMEOUT_SEC", "20"))
REQUEST_INTERVAL_SEC = float(os.getenv("REQUEST_INTERVAL_SEC", "1.0"))
MAX_FETCH_RETRIES = int(os.getenv("MAX_FETCH_RETRIES", "0"))
USER_AGENT = os.getenv(
    "MONITOR_USER_AGENT",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
)

_BROWSER_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

NEGATIVE_STOCK_SIGNALS = [
    "在庫なし",
    "完売御礼",
    "完売",
    "販売終了",
    "ただいま購入できません",
    "売り切れ",
    "品切れ",
    "在庫切れ",
    "入荷待ち",
    "再入荷をお知らせ",
    "現在お取り扱いできません",
]
POSITIVE_STOCK_SIGNALS = [
    "カートに入れる",
    "ご注文手続き",
    "ご購入手続き",
    "購入手続き",
    "即納（在庫あり）",
    "在庫あり",
    "注文する",
    "今すぐ購入",
]
JS_REQUIRED_SIGNALS = [
    "JavaScriptを有効にする必要があります",
    "このページではjavascriptを使用しています",
    "javascriptを使用しています",
    "JavaScript",
]
DISCOVERY_HINTS = {"search_discovery", "manual", "manual_check"}


@dataclass
class Product:
    jan: str
    product_name: str
    enoking_buy_price_yen: int
    required_condition: str
    buyback_source: str = ""


@dataclass
class Supplier:
    jan: str
    supplier: str
    url: str
    expected_price_yen: int | None
    shipping_included: bool
    condition_required: str
    enabled: bool
    parser_hint: str
    notes: str
    restock_watch: bool = False


@dataclass
class ParseResult:
    parsed_price_yen: int | None
    in_stock: bool | None
    raw_signals: str
    notes: str


def parse_bool(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def parse_int(value: str | None) -> int | None:
    value = str(value or "").strip().replace(",", "")
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        return None


def load_products(path: Path = CONFIG_DIR / "products_sample.csv") -> dict[str, Product]:
    products: dict[str, Product] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if not row or not row.get("jan"):
                continue
            product = Product(
                jan=row["jan"].strip(),
                product_name=row.get("product_name", "").strip(),
                enoking_buy_price_yen=int(row.get("enoking_buy_price_yen") or 0),
                required_condition=row.get("required_condition", "new").strip(),
                buyback_source=row.get("buyback_source", "").strip(),
            )
            products[product.jan] = product
    return products


def load_suppliers(path: Path = CONFIG_DIR / "supplier_urls.csv") -> list[Supplier]:
    suppliers: list[Supplier] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if not row or not row.get("jan") or not row.get("url"):
                continue
            suppliers.append(
                Supplier(
                    jan=row["jan"].strip(),
                    supplier=row.get("supplier", "").strip(),
                    url=row["url"].strip(),
                    expected_price_yen=parse_int(row.get("expected_price_yen")),
                    shipping_included=parse_bool(row.get("shipping_included")),
                    condition_required=row.get("condition_required", "new").strip(),
                    enabled=parse_bool(row.get("enabled", "true")),
                    parser_hint=row.get("parser_hint", "generic").strip(),
                    notes=row.get("notes", "").strip(),
                    restock_watch=parse_bool(row.get("restock_watch", "false")),
                )
            )
    return suppliers


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text.replace("\u3000", " ")).strip()


def yen_to_int(value: str) -> int | None:
    return parse_int(value.replace("円", "").replace("税込", ""))


def extract_prices(text: str) -> list[int]:
    prices: list[int] = []
    pattern = r"(?:￥|価格[:：]?|税込|本体価格)?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]{4,6})\s*円?"
    for match in re.finditer(pattern, text):
        price = yen_to_int(match.group(1))
        if price and 10_000 <= price <= 200_000:
            prices.append(price)
    return prices


def first_price_after_label(text: str, label_pattern: str) -> int | None:
    match = re.search(
        label_pattern + r".{0,140}?([0-9]{1,3}(?:,[0-9]{3})+|[0-9]{4,6})\s*円",
        text,
        re.DOTALL,
    )
    if match:
        return yen_to_int(match.group(1))
    return None


def stock_signal_summary(text: str) -> str:
    signals: list[str] = []
    for signal in NEGATIVE_STOCK_SIGNALS + POSITIVE_STOCK_SIGNALS + JS_REQUIRED_SIGNALS:
        if signal in text:
            signals.append(signal)
    return "|".join(dict.fromkeys(signals))


def detect_stock(text: str) -> bool | None:
    if any(signal in text for signal in NEGATIVE_STOCK_SIGNALS):
        return False
    if any(signal in text for signal in POSITIVE_STOCK_SIGNALS):
        return True
    return None


def parse_yahoo(text: str) -> ParseResult:
    return ParseResult(first_price_after_label(text, r"価格"), detect_stock(text), stock_signal_summary(text), "")


def parse_nojima(text: str) -> ParseResult:
    prices = extract_prices(text)
    price = min(prices) if prices else None
    return ParseResult(price, detect_stock(text), stock_signal_summary(text), "")


def parse_aeon(text: str) -> ParseResult:
    price = first_price_after_label(text, r"税込") or first_price_after_label(text, r"本体価格")
    return ParseResult(price, detect_stock(text), stock_signal_summary(text), "")


def parse_yodobashi(text: str) -> ParseResult:
    match = re.search(r"￥\s*([0-9]{1,3}(?:,[0-9]{3})+)", text)
    price = yen_to_int(match.group(1)) if match else None
    if price is None:
        price = first_price_after_label(text, r"価格")
    return ParseResult(price, detect_stock(text), stock_signal_summary(text), "")


def parse_search_discovery(text: str) -> ParseResult:
    normalized = normalize_text(text)
    notes = ["DISCOVERY_ONLY_SEARCH_PAGE", "NO_BUY_DECISION"]
    if any(signal in normalized for signal in JS_REQUIRED_SIGNALS):
        notes.append("NEEDS_BROWSER_OR_MANUAL_CHECK")
    return ParseResult(None, None, stock_signal_summary(normalized), "|".join(notes))


def parse_generic(text: str) -> ParseResult:
    prices = extract_prices(text)
    return ParseResult(min(prices) if prices else None, detect_stock(text), stock_signal_summary(text), "")


def parse_stock_only(text: str) -> ParseResult:
    """Stock signal only; price is taken from config expected_price_yen.

    Used for pages where requests-only price extraction is unreliable (login-walled
    Costco, JS-rendered Sony Store) but the static HTML still carries stock wording
    such as 在庫切れ / 入荷待ち / カートに入れる. Avoids mis-reading model numbers
    (e.g. CFIJ-10018) as prices.
    """
    return ParseResult(None, detect_stock(text), stock_signal_summary(text), "STOCK_ONLY_PRICE_FROM_CONFIG")


def parse_page(text: str, supplier: Supplier) -> ParseResult:
    normalized = normalize_text(text)
    hint = supplier.parser_hint.lower()
    if hint in DISCOVERY_HINTS:
        return parse_search_discovery(normalized)
    if hint == "stock_only":
        return parse_stock_only(normalized)
    if any(signal in normalized for signal in JS_REQUIRED_SIGNALS):
        return ParseResult(None, None, stock_signal_summary(normalized), "NEEDS_BROWSER_OR_MANUAL_CHECK")
    if hint == "yahoo":
        return parse_yahoo(normalized)
    if hint == "nojima":
        return parse_nojima(normalized)
    if hint == "aeon":
        return parse_aeon(normalized)
    if hint == "yodobashi":
        return parse_yodobashi(normalized)
    return parse_generic(normalized)


def is_discovery_only(supplier: Supplier, parsed: ParseResult | None = None) -> bool:
    return supplier.parser_hint.lower() in DISCOVERY_HINTS or bool(parsed and "DISCOVERY_ONLY" in parsed.notes)


def fetch(url: str, timeout: int | None = None, retries: int | None = None) -> tuple[int | None, str, str]:
    if not url.startswith(("http://", "https://")):
        return None, "", f"FETCH_SKIPPED: unsupported URL scheme: {url}"
    headers = {**_BROWSER_HEADERS, "User-Agent": USER_AGENT}
    effective_timeout = timeout if timeout is not None else REQUEST_TIMEOUT_SEC
    effective_retries = retries if retries is not None else MAX_FETCH_RETRIES
    last_error = ""
    for attempt in range(effective_retries + 1):
        if attempt > 0:
            time.sleep(3)
        try:
            response = requests.get(url, headers=headers, timeout=effective_timeout)
            response.encoding = response.apparent_encoding or response.encoding
            return response.status_code, response.text, ""
        except requests.Timeout as exc:
            last_error = f"FETCH_ERROR: {exc.__class__.__name__}: {exc}"
        except requests.RequestException as exc:
            return None, "", f"FETCH_ERROR: {exc.__class__.__name__}: {exc}"
    return None, "", last_error


def evaluate(product: Product, supplier: Supplier, result: ParseResult) -> dict[str, Any]:
    if is_discovery_only(supplier, result):
        return {
            "effective_price_yen": None,
            "gross_profit_yen": None,
            "is_buy_candidate": False,
            "decision_reason": "DISCOVERY_ONLY_NO_BUY_DECISION",
        }
    price = result.parsed_price_yen or supplier.expected_price_yen
    gross_profit = product.enoking_buy_price_yen - price if price else None
    buy_candidate = (
        price is not None
        and result.in_stock is True
        and supplier.shipping_included is True
        and supplier.condition_required == product.required_condition == "new"
        and gross_profit is not None
        and gross_profit >= BUY_MARGIN_THRESHOLD_YEN
    )
    reason = "BUY_CANDIDATE" if buy_candidate else "NOT_BUY_CANDIDATE"
    if price is None:
        reason = "NO_PRICE"
    elif result.in_stock is not True:
        reason = "NO_POSITIVE_STOCK_SIGNAL"
    elif gross_profit is not None and gross_profit < BUY_MARGIN_THRESHOLD_YEN:
        reason = "MARGIN_BELOW_THRESHOLD"
    return {
        "effective_price_yen": price,
        "gross_profit_yen": gross_profit,
        "is_buy_candidate": buy_candidate,
        "decision_reason": reason,
    }


def build_row(product: Product, supplier: Supplier, checked_at: str) -> dict[str, Any]:
    is_discovery = supplier.parser_hint.lower() in DISCOVERY_HINTS
    timeout = 10 if is_discovery else REQUEST_TIMEOUT_SEC
    retries = 0 if is_discovery else MAX_FETCH_RETRIES
    status_code, html, fetch_error = fetch(supplier.url, timeout=timeout, retries=retries)
    if fetch_error and "Timeout" in fetch_error:
        fetch_error = fetch_error + "|TIMEOUT_BLOCKED"
    parsed = ParseResult(None, None, "", fetch_error) if fetch_error else parse_page(html, supplier)
    eval_result = evaluate(product, supplier, parsed)
    return {
        "checked_at_jst": checked_at,
        "jan": supplier.jan,
        "product_name": product.product_name,
        "supplier": supplier.supplier,
        "url": supplier.url,
        "http_status": status_code,
        "fetch_ok": not bool(fetch_error),
        "parser_hint": supplier.parser_hint,
        "parsed_price_yen": parsed.parsed_price_yen,
        "expected_price_yen": supplier.expected_price_yen,
        "effective_price_yen": eval_result["effective_price_yen"],
        "enoking_buy_price_yen": product.enoking_buy_price_yen,
        "buyback_source": product.buyback_source,
        "gross_profit_yen": eval_result["gross_profit_yen"],
        "in_stock": parsed.in_stock,
        "restock_watch": supplier.restock_watch,
        "shipping_included": supplier.shipping_included,
        "condition_required": supplier.condition_required,
        "is_buy_candidate": eval_result["is_buy_candidate"],
        "decision_reason": eval_result["decision_reason"],
        "raw_signals": parsed.raw_signals,
        "notes": "; ".join(part for part in [supplier.notes, parsed.notes] if part),
    }


def dispatch_message(body: str) -> None:
    """Send one notification body to all configured channels (Telegram/Slack/Discord)."""
    plain_body = re.sub(r"<[^>]+>", "", body)
    print(f"::notice::{plain_body}")

    telegram_token = os.getenv("TELEGRAM_BOT_TOKEN")
    telegram_chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if telegram_token and telegram_chat_id:
        try:
            requests.post(
                f"https://api.telegram.org/bot{telegram_token}/sendMessage",
                json={"chat_id": telegram_chat_id, "text": body, "parse_mode": "HTML"},
                timeout=10,
            )
        except requests.RequestException as exc:
            print(f"telegram notification failed: {exc}", file=sys.stderr)

    for env_name, payload_key in [("SLACK_WEBHOOK_URL", "text"), ("DISCORD_WEBHOOK_URL", "content")]:
        webhook = os.getenv(env_name)
        if webhook:
            try:
                requests.post(webhook, json={payload_key: plain_body}, timeout=10)
            except requests.RequestException as exc:
                print(f"notification failed ({env_name}): {exc}", file=sys.stderr)


def notify(candidates: list[dict[str, Any]]) -> None:
    if not candidates:
        return
    lines = ["🛒 <b>Enoking転売候補が見つかりました！</b>"]
    for row in candidates[:10]:
        lines.append(
            f"📦 {row['product_name']}\n"
            f"  🏪 {row['supplier']}  💴 {row['effective_price_yen']:,}円\n"
            f"  💰 粗利: {row['gross_profit_yen']:,}円（売り先: {row.get('buyback_source') or 'エノキング'} {row['enoking_buy_price_yen']:,}円）\n"
            f"  🔗 {row['url']}"
        )
    dispatch_message("\n\n".join(lines))


def notify_restocks(restocks: list[dict[str, Any]]) -> None:
    """Distinct alert for out-of-stock -> in-stock transitions on watched suppliers."""
    if not restocks:
        return
    lines = ["🔔 <b>入荷検知！（在庫切れ→在庫あり）</b>"]
    for row in restocks[:10]:
        price = row.get("effective_price_yen")
        gross = row.get("gross_profit_yen")
        price_text = f"{price:,}円" if isinstance(price, int) else "価格は要確認"
        margin_text = f"／想定粗利 {gross:,}円" if isinstance(gross, int) else ""
        lines.append(
            f"📦 {row['product_name']}\n"
            f"  🏪 {row['supplier']}  💴 {price_text}{margin_text}\n"
            f"  💰 売り先: {row.get('buyback_source') or 'エノキング'} {row['enoking_buy_price_yen']:,}円\n"
            f"  🔗 {row['url']}"
        )
    dispatch_message("\n\n".join(lines))


DB_PATH = OUTPUT_DIR / "resale_db.sqlite"

_CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS monitor_results (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    checked_at_jst TEXT NOT NULL,
    jan TEXT NOT NULL,
    product_name TEXT,
    supplier TEXT,
    url TEXT,
    http_status INTEGER,
    fetch_ok INTEGER,
    parser_hint TEXT,
    parsed_price_yen INTEGER,
    expected_price_yen INTEGER,
    effective_price_yen INTEGER,
    enoking_buy_price_yen INTEGER,
    gross_profit_yen INTEGER,
    in_stock INTEGER,
    shipping_included INTEGER,
    condition_required TEXT,
    is_buy_candidate INTEGER,
    decision_reason TEXT,
    raw_signals TEXT,
    notes TEXT,
    UNIQUE(checked_at_jst, jan, supplier)
)
"""


def write_db(rows: list[dict[str, Any]], db_path: Path = DB_PATH) -> int:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(_CREATE_TABLE_SQL)
        inserted = 0
        for row in rows:
            try:
                conn.execute(
                    """INSERT OR IGNORE INTO monitor_results
                    (checked_at_jst, jan, product_name, supplier, url, http_status,
                     fetch_ok, parser_hint, parsed_price_yen, expected_price_yen,
                     effective_price_yen, enoking_buy_price_yen, gross_profit_yen,
                     in_stock, shipping_included, condition_required,
                     is_buy_candidate, decision_reason, raw_signals, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        row.get("checked_at_jst"),
                        row.get("jan"),
                        row.get("product_name"),
                        row.get("supplier"),
                        row.get("url"),
                        row.get("http_status"),
                        int(bool(row.get("fetch_ok"))),
                        row.get("parser_hint"),
                        row.get("parsed_price_yen"),
                        row.get("expected_price_yen"),
                        row.get("effective_price_yen"),
                        row.get("enoking_buy_price_yen"),
                        row.get("gross_profit_yen"),
                        None if row.get("in_stock") is None else int(bool(row.get("in_stock"))),
                        int(bool(row.get("shipping_included"))),
                        row.get("condition_required"),
                        int(bool(row.get("is_buy_candidate"))),
                        row.get("decision_reason"),
                        row.get("raw_signals"),
                        row.get("notes"),
                    ),
                )
                inserted += conn.execute("SELECT changes()").fetchone()[0]
            except sqlite3.Error as exc:
                print(f"db write error for {row.get('jan')}/{row.get('supplier')}: {exc}", file=sys.stderr)
        conn.commit()
    return inserted


def autosize_worksheet(ws: Any) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in ws.columns:
        col_letter = get_column_letter(column[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 70)


def write_excel(rows: list[dict[str, Any]], summary: dict[str, Any], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["metric", "value"])
    for key, value in summary.items():
        ws.append([key, value])
    autosize_worksheet(ws)

    fieldnames = list(rows[0].keys()) if rows else ["message"]
    result_sheets = {
        "Results": rows,
        "BuyCandidates": [row for row in rows if row.get("is_buy_candidate")],
        "DiscoveryOnly": [row for row in rows if row.get("decision_reason") == "DISCOVERY_ONLY_NO_BUY_DECISION"],
    }
    for sheet_name, sheet_rows in result_sheets.items():
        ws2 = wb.create_sheet(sheet_name)
        ws2.append(fieldnames)
        if sheet_rows:
            for row in sheet_rows:
                ws2.append([row.get(field) for field in fieldnames])
        elif sheet_name == "BuyCandidates":
            ws2.append(["NO_BUY_CANDIDATE"] + [None] * (len(fieldnames) - 1))
        autosize_worksheet(ws2)
    wb.save(out_path)


def write_outputs(rows: list[dict[str, Any]], summary: dict[str, Any], stamp: str) -> tuple[Path, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_csv = OUTPUT_DIR / f"monitor_result_{stamp}.csv"
    out_xlsx = OUTPUT_DIR / f"monitor_result_{stamp}.xlsx"
    fieldnames = list(rows[0].keys()) if rows else ["message"]
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        if rows:
            writer.writerows(rows)
        else:
            writer.writerow({"message": "NO_ROWS"})
    write_excel(rows, summary, out_xlsx)
    latest_csv = OUTPUT_DIR / "latest.csv"
    latest_xlsx = OUTPUT_DIR / "latest.xlsx"
    shutil.copyfile(out_csv, latest_csv)
    shutil.copyfile(out_xlsx, latest_xlsx)
    return out_csv, out_xlsx


def previous_in_stock(jan: str, supplier: str, db_path: Path = DB_PATH) -> int | None:
    """Most recent prior in_stock value (0/1) for a jan+supplier, or None if never recorded.

    Must be called BEFORE the current run's rows are written to the DB.
    """
    if not db_path.exists():
        return None
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.execute(
                "SELECT in_stock FROM monitor_results "
                "WHERE jan = ? AND supplier = ? AND in_stock IS NOT NULL "
                "ORDER BY id DESC LIMIT 1",
                (jan, supplier),
            )
            row = cursor.fetchone()
    except sqlite3.Error as exc:
        print(f"restock lookup error for {jan}/{supplier}: {exc}", file=sys.stderr)
        return None
    return None if row is None else row[0]


def detect_restocks(rows: list[dict[str, Any]], db_path: Path = DB_PATH) -> list[dict[str, Any]]:
    """Return rows that just transitioned from out-of-stock to in-stock.

    Conservative: fires only when the current check shows in_stock True and the
    most recent prior recorded state for the same (jan, supplier) was explicitly
    out-of-stock (0). First-ever observations do not fire to avoid launch noise.
    """
    events: list[dict[str, Any]] = []
    for row in rows:
        if row.get("in_stock") is not True:
            continue
        if previous_in_stock(row["jan"], row["supplier"], db_path) == 0:
            events.append(row)
    return events


def main() -> int:
    now = datetime.now(JST)
    checked_at = now.strftime("%Y-%m-%d %H:%M:%S %Z")
    stamp = now.strftime("%Y%m%d_%H%M%S")
    products = load_products()
    suppliers = [supplier for supplier in load_suppliers() if supplier.enabled]
    if parse_bool(os.getenv("RESTOCK_WATCH_ONLY")):
        suppliers = [supplier for supplier in suppliers if supplier.restock_watch]
    rows: list[dict[str, Any]] = []

    for idx, supplier in enumerate(suppliers):
        product = products.get(supplier.jan)
        if not product:
            print(f"Skipping unknown JAN: {supplier.jan}", file=sys.stderr)
            continue
        if idx > 0 and REQUEST_INTERVAL_SEC > 0:
            time.sleep(REQUEST_INTERVAL_SEC)
        rows.append(build_row(product, supplier, checked_at))

    candidates = [row for row in rows if row.get("is_buy_candidate")]
    # Detect restocks BEFORE writing this run's rows (compares against prior DB state).
    restocks = detect_restocks(rows)
    summary: dict[str, Any] = {
        "checked_at_jst": checked_at,
        "checked": len(rows),
        "buy_candidates": len(candidates),
        "restocks_detected": len(restocks),
        "discovery_only": sum(1 for row in rows if row.get("decision_reason") == "DISCOVERY_ONLY_NO_BUY_DECISION"),
        "fetch_errors": sum(
            1 for row in rows
            if not row.get("fetch_ok") and row.get("parser_hint", "") not in DISCOVERY_HINTS
        ),
        "fetch_errors_discovery": sum(
            1 for row in rows
            if not row.get("fetch_ok") and row.get("parser_hint", "") in DISCOVERY_HINTS
        ),
    }
    out_csv, out_xlsx = write_outputs(rows, summary, stamp)
    summary["output_csv"] = str(out_csv)
    summary["output_xlsx"] = str(out_xlsx)
    summary["latest_csv"] = str(OUTPUT_DIR / "latest.csv")
    summary["latest_xlsx"] = str(OUTPUT_DIR / "latest.xlsx")
    write_excel(rows, summary, out_xlsx)
    shutil.copyfile(out_xlsx, OUTPUT_DIR / "latest.xlsx")
    inserted = write_db(rows)
    summary["db_rows_inserted"] = inserted
    summary["db_path"] = str(DB_PATH)
    notify(candidates)
    notify_restocks(restocks)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
